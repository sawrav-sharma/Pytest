from dataclasses import dataclass
from numpy import column_stack
from openpyxl import Workbook
import openpyxl
import csv

def getRowCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def getColumnCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

def getColumnCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)


